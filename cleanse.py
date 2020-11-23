# Cleanse tool
# Aithor: Aidan Hall
# November 2020 
# https://github.com/aidanHallLondon/CZD/blob/master/cleanse.py
#
# - loads data from a specific Spreadsheet file 
# (built in a very specific way) processes it and generates a new spreadsheet.
# 
# Takes the ReadingData spreadsheet and adds binary columns for some dimensions (Columns).
# For those columns it adds a new column for all unique values and sets to 1/0 if there is a match
# Also adds a meta data sheet to help debug  the  data and code
#
import re
import sys
from datetime import datetime

import openpyxl as op
from openpyxl.utils import get_column_letter

from spreadsheetStructure import *
from xlsxFormats import getFormats

# import pprint
# import operator

columns = createColumnData()

if sys.version_info[0] < 3:
    raise Exception(
        f"Python version : {sys.version_info}\n\nYou need to run this using Python 3 - check your configuration or use the Python3 command")
#
# You can speed up the program and the loading of the spreadsheet for testing by only processing 100 rows
# BE VERY CAREFUL
limitOutputToFirst100 = False
# now = datetime.utcnow()
# pp = pprint.PrettyPrinter(indent=4)

sourceFileName = 'Reading data to upload.xlsx'
outputFileName = "data.xlsx"
# name of the sheet we wnat in the workbook
mainSheetName = "coded"
# members contain lots of spaces and messy chars that mess up heading names
# this regEx is used to remove them
shortnameRegExpPattern = "[^|a-zA-Z0-9]+"
# this is rounding a float before multiplying by 100 to get %age
percentageRoundingPlaces = 6
# width of columns with 0 or 1 boolean data in them
# small sizes make the sheet more manageable but the headings harder to read
booleanColumnWidth = 3


#
# These stlyes are used for formatting the output documents
#


formats=getFormats()

def main():
    sourceXL = op.load_workbook(
        sourceFileName, read_only=True, data_only='True')
    dataXL = op.Workbook()
    dataWs = dataXL.active
    metaWs = dataXL.create_sheet('Meta data')

    # wb = load_workbook(sourceFileName)
    # grab the active worksheet
    ws = sourceXL.active
    if mainSheetName in sourceXL.sheetnames:
        ws = sourceXL[mainSheetName]
    else:
        raise Exception(
            f"The sheet called '{mainSheetName}' isn't there, check you gave not remnamed it or got the wrong spreadsheet.")

    # go through our column list and lookup the position of each column Name
    # in the workSheet
    getSourceColumPositions(ws)

    # get list of all unique members in each column (as marked)
    getColumnMembers(ws)  # sets columnWidth and [members]

    # set output column positions based on width of all columns including new ones
    # bool columns are added for each member found
    computeOutputColumnWidths()
    computeOutputColumnPositions()

    outputMetaData(ws, metaWs)

    # Output header row and data rows
    outputHeadingRow(dataWs, 1)
    outputcolumnsByRow(ws, dataWs, 2, limitOutputToFirst100)
    if limitOutputToFirst100:
        print('DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER ')
        print('The output has been limited to 100 items only change limitOutputToFirst100=True to limitOutputToFirst100=False ')
        print('DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER DANGER ')
    dataWs.freeze_panes = 'B2'
    metaWs.freeze_panes = 'B2'
    # Save to the dataXL file
    dataXL.save(outputFileName)
    # pp.pprint(columns)


def getColumnHeadingNames(ws):
    rows = ws.iter_rows(min_row=1, max_row=1)  # returns a generator of rows
    first_row = next(rows)  # get the first row
    # extract the values from the cells
    headings = [colItem.value for colItem in first_row]
    return headings


def getSourceColumPositions(ws):
    # go through our column list and lookup the position of each column Name
    # in the workSheet
    headings = getColumnHeadingNames(ws)
    #
    for colItem in columns:
        colName = colItem['name']
        colItem['outputName'] = colItem.get(
            'outputName', getShortName(colName))
        #
        try:
            colItem['srcIndex'] = headings.index(colName)
        except ValueError:
            errMsg = f"ERROR column not found : {colName}"
            colItem['srcIndex'] = -1
            colItem['Errors'] = colItem.get('Errors', []).append(errMsg)
            print(errMsg)


def getColumnMembers(ws):
    # scan every column that has options, i.e. a limited set of members and
    # make a dict of them with counts
    def getMemberCount(member):
        return member['count']
    #
    idIndex = next((col for col in columns if col['name'] == 'ID'))['srcIndex']
    validIndex = next((col for col in columns if col['name'] == 'Valid'))[
        'srcIndex']
    # add memberDict nodes

    for colItem in columns:
        getMembersMode = colItem.get('getMembers', 'none')
        if getMembersMode != 'none':
            colItem['memberDict'] = {}  # empty set
            colItem['memberCount'] = 0  # force it to zero
        colItem['includeInvalidRows'] = colItem.get(
            'includeInvalidRows', False)
        colItem['rowCount'] = 0  # force it to zero
        colItem['invalidRowCount'] = 0  # force it to zero
    # get members
    rowIndex = 2
    for row in ws.iter_rows(min_row=rowIndex):
        # ignore any row without an ID
        if row[idIndex].value and int(row[idIndex].value) > 0:
            for colItem in columns:
                # discount any row that is not explicitly Valid
                try:
                    isValidRow = int(float(row[validIndex].value)) == 1
                except:
                    isValidRow = False
                if isValidRow or colItem['includeInvalidRows']:
                    getMembersMode = colItem.get('getMembers', 'none')
                    if getMembersMode != 'none':
                        members = colItem.get('memberDict', [])
                        # Get the raw value Check for type rules and apply
                        srcCell = row[colItem["srcIndex"]]
                        rawValue = str(srcCell.value)
                        if getMembersMode == 'members':
                            members = colItem.get('memberDict', [])
                            # Get the raw value Check for type rules and apply
                            srcCell = row[colItem["srcIndex"]]
                            rawValue = str(srcCell.value)
                            # TODO can we curry the value after building the list and only run this for the UNIQUE items not all 5000 in every column?
                            value = curryValueByColumnItem(rawValue, colItem)
                            # add count to memberslist if not already there
                            if not(value in members):
                                members[value] = {
                                    'count': 0,
                                    'shortName': getShortName(value)
                                }
                            # increment the count
                            members[value]['count'] += 1
                            # TODO why is this different to the (sum of all) members list counts
                            if rawValue is not None and rawValue != '':
                                colItem['memberCount'] += 1
                        if getMembersMode == 'tokens':
                            # TODO can we curry the value after building the list and only run this for the UNIQUE items not all 5000 in every column?
                            values = rawValue.split('|')
                            # add count to memberslist if not already there
                            for value in values:
                                if not(value in members):
                                    members[value] = {
                                        'count': 0,
                                        'shortName': getShortName(value)
                                    }
                                # increment the count
                                members[value]['count'] += 1
                                # TODO why is this different to the (sum of all) members list counts
                                if value is not None and value != '':
                                    colItem['memberCount'] += 1
                else:
                    colItem['invalidRowCount'] += 1
                colItem['rowCount'] += 1
        rowIndex += 1

    # after collecting the members summarise them and discard excess tokens
    for colItem in columns:
        getMembersMode = colItem.get('getMembers', 'none')
        if getMembersMode != 'none':
            colItem['memberPercentage'] = round(
                colItem['memberCount']/colItem['rowCount'], percentageRoundingPlaces)
            if getMembersMode != 'none':
                # sort members list by count DESC
                membersArray = []
                memberDict = colItem.get('memberDict', [])
                # count all  the tokens
                totalMemberCount=0
                for key in memberDict.keys():
                    totalMemberCount += memberDict[key]['count']
                # add percentages and copy  to the new Array for sorting and culling
                for key in memberDict.keys():
                    memberDict[key]['name'] = key
                    memberDict[key]['percentage'] = round(memberDict[key]['count']/totalMemberCount, percentageRoundingPlaces)
                    membersArray.append(memberDict[key])
                # Sort
                membersArray = sorted(membersArray, key=getMemberCount, reverse=True)
                # remove excess members
                if getMembersMode == 'tokens':
                    selectTopTokensDefault=selectTopTokens
                if getMembersMode == 'members':
                    selectTopTokensDefault=0
                selectTopTokens = colItem.get('selectTopTokens', selectTopTokensDefault)
                if selectTopTokens > 0:
                    membersArray=membersArray[:selectTopTokens]
                # convert to a dict again
                # TODO Maybe it is better as an array for all of these not just tokens
                memberDict = {}
                for member in membersArray:
                    memberDict[member['name']] = member
                #print(memberDict)
                colItem['memberDict'] = memberDict


def getShortName(value):
    # convert a string into a simpler shorter name (alphanumeric and bar only)
    return re.sub(shortnameRegExpPattern, "", str(value))


def curryValueByColumnItem(value, colItem):
    typeCode = colItem.get('type', 'string')
    #
    if typeCode == 'string':
        return str(value)
    #
    if typeCode == 'integer':
        try:
            value2 = str(int(float(value)))
        except:
            value2 = 'NULL'
            print(
                f"FAIL integer type converiosn attempted on {colItem['name']}.{value}")
        value = value2
    #
    if typeCode == 'head':
        delimeter = colItem.get('delimeter', '|')
        value = (((value or '')+delimeter).split(delimeter)[0]).strip()
    #
    return value


def computeOutputColumnPositions():
    # set output column positions based on width of all columns including new ones
    # bool columns are added for each member found
    outputIndex = 1
    for colItem in columns:
        if colItem.get('output', True):
            colItem['outputIndex'] = outputIndex
            # increment for next column
            outputIndex = outputIndex + colItem['columnWidth']


def computeOutputColumnWidths():
    # how wide is a column (1) and all the new cols it generates
    # how many cols does it take up in the output
    for colItem in columns:
        if colItem.get('output', True):
            columnWidth = 1
            if 'memberDict' in colItem:
                # the width is main dcolumn + one for each member in boolean cols
                columnWidth = len(colItem['memberDict']) + 1
        else:
            columnWidth = 0
        colItem['columnWidth'] = columnWidth


def outputHeadingRow(dataWs, startRowIndex):
    rowIndex = startRowIndex
    dataWs.column_dimensions['A'].width = 7
    for colItem in columns:
        if colItem.get('output', True):
            outputName = colItem["outputName"]
            setCell(dataWs, rowIndex,
                    colItem["outputIndex"], outputName, formats['header'])
            #
            # Generate boolean member column headings
            if colItem.get('getMembers', 'none') != 'none':
                colIndex = colItem["outputIndex"]+1
                for key in sorted(colItem['memberDict'].keys()):
                    m = colItem['memberDict'][key]
                    # The get function seems to use the default for the empty  string valeu which we need to capture so an if statement is used
                    if 'memberColumnSuffix' in colItem:
                        memberColumnSuffix = colItem['memberColumnSuffix']
                    else:
                        memberColumnSuffix = ('_' + outputName)
                    columnHeadingName = f"{m['shortName']}{memberColumnSuffix}"
                    # if colItem['name']=='Under72incInUtero':
                    #    print(f'{columnHeadingName}')
                    #
                    setCell(dataWs,
                            rowIndex, colIndex,
                            columnHeadingName,
                            formats['memberHeader']
                            )
                    colLetter = get_column_letter(colIndex)
                    dataWs.column_dimensions[colLetter].width = booleanColumnWidth
                    colIndex = colIndex + 1
    rowIndex = rowIndex+1
    return rowIndex


def setCell(dataWs, rowIndex, colIndex, value, style):
    cell = dataWs.cell(rowIndex, colIndex)
    cell.value = value
    if ('border' in style):
        cell.border = style['border']
    if ('font' in style):
        cell.font = style['font']
    if ('number_format' in style):
        cell.number_format = style['number_format']
    return cell


def outputcolumnsByRow(ws, dataWs, startRowIndex, limitedTo100=False):
    rowIndex = startRowIndex
    idIndex = next((col for col in columns if col['name'] == 'ID'))['srcIndex']
    validIndex = next((col for col in columns if col['name'] == 'Valid'))[
        'srcIndex']
    #
    max_row = 100 if limitedTo100 else 65530
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        try:
            isValidRow = int(float(row[validIndex].value)) == 1
        except:
            isValidRow = False
        if row[idIndex].value and int(row[idIndex].value) > 0 and isValidRow:
            for colItem in columns:
                #selectTopTokens = colItem.get('getMembersMode', 10)
                delimeter = colItem.get('delimeter', '|')
                if colItem.get('output', True):
                    # Get the value and convert it's type as requested
                    value = curryValueByColumnItem(
                        row[colItem["srcIndex"]].value, colItem)
                    dataWs.cell(
                        row=rowIndex,
                        column=colItem["outputIndex"]
                    ).value = value
                    # generate boolean member column data
                    getMembersMode = colItem.get('getMembers', 'none')
                    if getMembersMode != 'none':
                        colIndex = colItem["outputIndex"]+1
                        members = colItem['memberDict']
                        for key in sorted(members.keys()):
                            # m = members[key]
                            # output a ONE or a ZERO based one which member column
                            # matches this source cell value
                            if getMembersMode == 'members':
                                booleanValue = str(
                                    1 if str(key) == str(value) else 0)
                            if getMembersMode == 'tokens':
                                booleanValue = str(1 if (
                                    str(delimeter + key + delimeter) in str(delimeter + value + delimeter)) else 0)
                            setCell(dataWs,
                                    rowIndex, colIndex,
                                    booleanValue,  formats['boolean' +
                                                           booleanValue]
                                    )
                            # formattingStyles['fonts'].booleanTrue
                            # next col
                            colIndex = colIndex + 1
            # next row
            rowIndex = rowIndex+1
            # print(rowIndex)
    print('Rows read')


def outputMetaData(ws, dataWs):
    colIndex = 1
    columnStepSize = 1
    columnWithMemberListStepSize = 5
    topRowIndex = 2
    rowIndex = 1

    metaValues = [
        {'name': 'name'},
        {'name': 'output'},
        {'name': 'outputName'},
        {'name': 'type'},
        {'name': 'srcIndex'},
        {'name': 'outputIndex'},
        {'name': 'getMembers'},
        {'name': 'selectTopTokens'},
        {'name': 'memberCount'},
        {'name': 'rowCount'},
        {'name': 'invalidRowCount'},
        {'name': 'memberPercentage'}

    ]

    # row labels
    rowIndex = topRowIndex
    dataWs.column_dimensions['A'].width = 26
    setCell(dataWs, 1, colIndex, 'Run', formats['header'])
    setCell(dataWs, 1, colIndex+1, str(datetime.utcnow()),
            formats['header'])
    for meta in metaValues:
        setCell(dataWs, rowIndex, colIndex,
                meta['name'], formats['rowHeader'])
        rowIndex += 1
    rowIndex += 1  # leave a gap
    setCell(dataWs, rowIndex, colIndex,
            'Members list', formats['rowHeader'])
    rowIndex += 1

    #
    # values for each Column in columns
    rowIndex = topRowIndex
    colIndex = colIndex + columnStepSize
    for colItem in columns:
        rowIndex = 2
        for meta in metaValues:
            dataWs.cell(rowIndex, colIndex).value = colItem.get(
                meta['name'], '')
            rowIndex += 1
        if 'memberDict' in colItem:
            memberDict = colItem['memberDict']
            totalCount = 0
            rowIndex += 1
            style = formats['header']
            if colItem.get('getMembers','members')=='members':
                setCell(dataWs, rowIndex, colIndex + 0, 'member', style)
                setCell(dataWs, rowIndex, colIndex + 1, 'shortName', style)
                setCell(dataWs, rowIndex, colIndex + 2, 'count', style)
                setCell(dataWs, rowIndex, colIndex + 3, 'percentage', style)
            if colItem.get('getMembers','members')=='tokens':
                setCell(dataWs, rowIndex, colIndex + 0, 'token', style)
                setCell(dataWs, rowIndex, colIndex + 1, 'count', style)
                setCell(dataWs, rowIndex, colIndex + 2, 'percentage', style)
            rowIndex += 1
            for key in memberDict.keys():
                member = memberDict[key]
                totalCount += member['count']

            percentageSum = 0
            for key in memberDict.keys():
                member = memberDict[key]
                count = member['count']
                percentage = memberDict[key]['percentage'] 
                if colItem.get('getMembers','members')=='members':
                    setCell(dataWs, rowIndex, colIndex + 0, key, formats['memberValue'])
                    setCell(dataWs, rowIndex, colIndex + 1,  member['shortName'], formats['memberValue'])
                    setCell(dataWs, rowIndex, colIndex + 2, count, formats['memberName'])
                    setCell(dataWs, rowIndex, colIndex + 3, percentage,  formats['percentage'])
                if colItem.get('getMembers','members')=='tokens':
                    setCell(dataWs, rowIndex, colIndex + 0, key, formats['memberValue'])
                    setCell(dataWs, rowIndex, colIndex + 1, count, formats['memberValue'])
                    setCell(dataWs, rowIndex, colIndex + 2, percentage,  formats['percentage'])
                percentageSum += percentage
                rowIndex += 1
            if colItem.get('getMembers','members')=='members':
                setCell(dataWs, rowIndex, colIndex + 3, percentageSum,  formats['percentage'])
            if colItem.get('getMembers','members')=='tokens':
                setCell(dataWs, rowIndex, colIndex + 2, percentageSum,  formats['percentage'])
            rowIndex += 1
            # thank you next
            colIndex = colIndex + columnWithMemberListStepSize
        else:
            # thank you next
            colIndex = colIndex + columnStepSize


if __name__ == '__main__':
    main()
